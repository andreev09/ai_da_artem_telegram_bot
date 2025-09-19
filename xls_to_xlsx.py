from __future__ import annotations

from io import BytesIO
from typing import ByteString

import xlrd
from openpyxl import Workbook


def convert_xls_bytes_to_xlsx_bytes(xls_bytes: ByteString) -> bytes:
    """Convert XLS file content (bytes) to XLSX bytes.

    This function reads the old BIFF .xls file with `xlrd` (requires xlrd==1.2.0)
    and writes an .xlsx file into memory using `openpyxl`.

    Args:
        xls_bytes: raw bytes of an .xls file

    Returns:
        bytes containing a .xlsx file
    """
    with BytesIO(xls_bytes) as inbuf:
        book = xlrd.open_workbook(file_contents=inbuf.read(), formatting_info=False)

    wb = Workbook()
    # remove default sheet created by Workbook
    default = wb.active
    wb.remove(default)

    for sheet_idx in range(book.nsheets):
        sheet = book.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=sheet.name[:31])  # Excel sheet name limit

        for r in range(sheet.nrows):
            row_vals = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                # xlrd returns different types; use value directly
                val = cell.value
                row_vals.append(val)
            # write row (openpyxl is 1-indexed)
            ws.append(row_vals)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

from io import BytesIO
import os
import shutil
import subprocess

import pytest
import xlwt
from openpyxl import load_workbook

from xls_to_xlsx import convert_xls_bytes_to_xlsx_bytes


def make_sample_xls_bytes() -> bytes:
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    ws.write(0, 0, "Name")
    ws.write(0, 1, "Age")
    ws.write(1, 0, "Alice")
    ws.write(1, 1, 30)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


@pytest.mark.skipif(
    shutil.which("soffice") is None and os.environ.get("SOFFICE_PATH") is None,
    reason="LibreOffice soffice binary is required for conversion tests.",
)
def test_convert_xls_to_xlsx_roundtrip():
    xls_bytes = make_sample_xls_bytes()
    xlsx_bytes = convert_xls_bytes_to_xlsx_bytes(xls_bytes)

    wb = load_workbook(filename=BytesIO(xlsx_bytes), read_only=True)
    assert "Sheet1" in wb.sheetnames
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "Name"
    assert rows[0][1] == "Age"
    assert rows[1][0] == "Alice"
    # note: xlwt writes numbers as floats sometimes -> allow numeric equality
    assert float(rows[1][1]) == 30.0


def test_convert_xls_to_xlsx_reports_failure(monkeypatch):
    xls_bytes = make_sample_xls_bytes()

    monkeypatch.setenv("SOFFICE_PATH", "/usr/bin/soffice")

    def fake_run(*args, **kwargs):
        command = kwargs.get("args", args[0] if args else None)
        raise subprocess.CalledProcessError(
            returncode=1,
            cmd=command,
            stderr=b"conversion failed",
        )

    monkeypatch.setattr("xls_to_xlsx.subprocess.run", fake_run)

    with pytest.raises(RuntimeError) as excinfo:
        convert_xls_bytes_to_xlsx_bytes(xls_bytes)

    assert "LibreOffice conversion failed" in str(excinfo.value)
    assert "conversion failed" in str(excinfo.value)
